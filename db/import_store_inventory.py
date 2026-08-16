import openpyxl, re, sys, os
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
SRC = sys.argv[1] if len(sys.argv) > 1 else os.environ.get('INVENTORY_EXCEL_PATH', 'STORE_INVENTORY.xlsx')

if not os.path.exists(SRC):
    print(f"Excel file not found at: {SRC}")
    print("Usage: python import_store_inventory.py <path_to_excel_file>")
    sys.exit(1)

wb = openpyxl.load_workbook(SRC, data_only=True)

MECH_EXPR = "(SELECT id FROM material_categories WHERE code='MECH')"
CLOTH_EXPR = "(SELECT id FROM material_categories WHERE code='CLOTH')"

rows_out = []  # (code, name, category_expr, uom, hsn_code, current_stock, unit_price, min_stock, bin_location)

def clean(v):
    return str(v).strip() if v is not None else None

def to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0

# Oil Seal: ITEM CODE, OIL SEAL(name), HSN CODE, RACK/BOX NO, PHY STOCK, RECIVED, BALANCE
ws = wb[' Oil Seal']
for r in ws.iter_rows(min_row=4, values_only=True):
    code, name, hsn, rack, phy = r[0], r[1], r[2], r[3], r[4]
    if not code: continue
    stock = to_num(phy)
    rows_out.append((clean(code), clean(name), MECH_EXPR, 'Nos', clean(hsn), stock, 0, max(1, round(stock*0.2)), clean(rack)))

# Bearing Final Sheet: S NO, ITEM CODE, ITEM WITH DETAIL, Phy Stock
ws = wb['Bearing Final Sheet']
for r in ws.iter_rows(min_row=4, values_only=True):
    _, code, detail, phy = r[0], r[1], r[2], r[3]
    if not code: continue
    stock = to_num(phy)
    rows_out.append((clean(code), clean(detail), MECH_EXPR, 'Nos', None, stock, 0, max(1, round(stock*0.2)), None))

# CLOTHING: S.No, ITEM CODE, ITEM NAME, HSN CODE, PHY QTY, RATE, IGST 5%, TOTAL AMOUNT
ws = wb['CLOTHING']
for r in ws.iter_rows(min_row=4, values_only=True):
    _, code, name, hsn, qty, rate, igst, total = r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7]
    if not code: continue
    stock = to_num(qty)
    rows_out.append((clean(code), clean(name), CLOTH_EXPR, 'Nos', clean(hsn), stock, to_num(rate), max(1, round(stock*0.2)), None))

# TYRE: S NO, ITEM CODE, ITEM WITH DETAIL, HSN CODE, PHY STOCK
ws = wb['TYRE']
for r in ws.iter_rows(min_row=4, values_only=True):
    _, code, detail, hsn, phy = r[0], r[1], r[2], r[3], r[4]
    if not code: continue
    stock = to_num(phy)
    rows_out.append((clean(code), clean(detail), MECH_EXPR, 'Nos', clean(hsn), stock, 0, max(1, round(stock*0.2)), None))

# TYRE COUPLING, PIN BUSH: S NO, ITEM CODE, ITEM WITH DETAIL, PHY STOCK
ws = wb['TYRE COUPLING, PIN BUSH']
for r in ws.iter_rows(min_row=4, values_only=True):
    _, code, detail, phy = r[0], r[1], r[2], r[3]
    if not code: continue
    stock = to_num(phy)
    rows_out.append((clean(code), clean(detail), MECH_EXPR, 'Nos', None, stock, 0, max(1, round(stock*0.2)), None))

# PUMP SLEEVE: S.NO, ITEM CODE, ITEM DETAILS, STOCK -PHY
ws = wb['PUMP SLEEVE']
for r in ws.iter_rows(min_row=4, values_only=True):
    _, code, detail, phy = r[0], r[1], r[2], r[3]
    if not code: continue
    stock = to_num(phy)
    rows_out.append((clean(code), clean(detail), MECH_EXPR, 'Nos', None, stock, 0, max(1, round(stock*0.2)), None))

out_path = BASE_DIR / 'seed_store_inventory_import.sql'
print(f"-- {len(rows_out)} materials parsed")
with open(out_path, 'w', encoding='utf-8') as f:
    f.write("-- Bulk import from STORE INVENTORY SOFT COPY.xlsx, 6 sheets (Oil Seal/Bearing/Clothing/Tyre/Tyre Coupling/Pump Sleeve).\n")
    f.write("BEGIN;\n")
    for code, name, cat, uom, hsn, stock, price, minstk, bin_loc in rows_out:
        name_e = name.replace("'", "''") if name else 'UNNAMED'
        hsn_e = f"'{hsn}'" if hsn else 'NULL'
        bin_e = f"'{bin_loc}'" if bin_loc else 'NULL'
        f.write(
            f"INSERT INTO materials (code,name,category_id,uom,hsn_code,current_stock,unit_price,min_stock,reorder_level,bin_location) "
            f"VALUES ('{code}','{name_e}',{cat},'{uom}',{hsn_e},{stock},{price},{minstk},{minstk},{bin_e}) "
            f"ON CONFLICT (code) DO NOTHING;\n"
        )
    f.write("COMMIT;\n")
print(f"wrote {out_path}")
